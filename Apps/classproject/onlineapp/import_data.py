
import openpyxl
import click
import os
import sys

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')

import django
django.setup()
from onlineapp.models import *

# c = College(name="Chaitanya Bharathi Institute Of technology", location="hyd", acronym="CBIT", contact="cbit@gmail.com")
# c.save()



def import_clg():
    workbook = openpyxl.load_workbook("students.xlsx")
    worksheet = workbook.get_sheet_by_name("Colleges")
    flag = 0
    for row in worksheet.rows:
        if not flag:
            flag = 1
            continue
        colleges = [cell.value for cell in row]
        c = College(name=colleges[0], location=colleges[2], acronym=colleges[1], contact=colleges[3])
        c.save()

#import_clg()
#student = Student()
def import_students():
    workbook = openpyxl.load_workbook("students.xlsx")
    worksheet = workbook.get_sheet_by_name("Current")
    flag = 0
    for row in worksheet.rows:
        if not flag:
            flag = 1
            continue
        student_values = [cell.value for cell in row]
        # student.name = student_values[0]
        # student.email = student_values[2]
        # student.db_folder = student_values[3]
        # student.college = College.objects.get(acronym=student_values[1])
        # student.save()
        student = Student(name =student_values[0],email= student_values[2], db_folder=student_values[3], college=College.objects.get(acronym=student_values[1]))
        student.save()
    worksheet = workbook.get_sheet_by_name("Deletions")
    flag = 0
    for row in worksheet.rows:
        if not flag:
            flag = 1
            continue
        student_values = [cell.value for cell in row]
        student = Student(name=student_values[0], email=student_values[2],dropped_out = True, db_folder=student_values[3],
                          college=College.objects.get(acronym=student_values[1]))
        student.save()
#import_students()

def import_marks():
    workbook = openpyxl.load_workbook("out.xlsx")
    worksheet = workbook.get_sheet_by_name("Test resul")
    flag = 0
    for row in worksheet.rows:
        if not flag:
            flag = 1
            continue
        marks_values = [cell.value for cell in row]
        mock = MockTest1(problem1 =marks_values[1],problem2= marks_values[2], problem3=marks_values[3], problem4=marks_values[4], total = marks_values[5], student = Student.objects.get(db_folder = marks_values[0].split("_")[-2]))
        mock.save()

import_marks()


